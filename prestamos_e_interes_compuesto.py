import io
from dataclasses import dataclass
from typing import Literal

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

Sistema = Literal["Francés", "Alemán"]
FORMATO_MONEDA = "$#,##0.00"
FILA_ENCABEZADO = 8


@dataclass(frozen=True)
class ConfiguracionPrestamo:
    nombre: str
    icono: str
    descripcion: str
    monto: float
    tasa: float
    plazo: int
    permite_elegir_sistema: bool = False


PRESTAMOS = {
    "🏠 Préstamo Hipotecario": ConfiguracionPrestamo(
        "Hipotecario", "🏠", "Calcula tu crédito de vivienda con el sistema francés.",
        50_000.0, 9.0, 20,
    ),
    "🏢 Préstamo Comercial": ConfiguracionPrestamo(
        "Comercial", "🏢", "Ideal para negocios y pymes.",
        15_000.0, 12.0, 5, True,
    ),
}


def calcular_amortizacion(
    monto: float, tasa_anual: float, anios: int, sistema: Sistema = "Francés"
) -> pd.DataFrame:
    """Genera una tabla de amortización francesa o alemana."""
    meses = anios * 12
    tasa_mensual = tasa_anual / 12
    saldo = monto
    filas = []

    if sistema == "Francés":
        cuota_fija = (
            monto / meses
            if tasa_mensual == 0
            else monto * tasa_mensual / (1 - (1 + tasa_mensual) ** -meses)
        )
        capital_fijo = None
    elif sistema == "Alemán":
        cuota_fija = 0.0
        capital_fijo = monto / meses
    else:
        raise ValueError(f"Sistema no válido: {sistema}")

    for mes in range(1, meses + 1):
        interes = saldo * tasa_mensual
        capital = cuota_fija - interes if sistema == "Francés" else capital_fijo
        cuota = cuota_fija if sistema == "Francés" else capital + interes

        # La última cuota absorbe pequeñas diferencias de punto flotante.
        if mes == meses:
            capital = saldo
            cuota = capital + interes

        saldo = max(0.0, saldo - capital)
        filas.append({
            "Mes": mes,
            "Cuota Mensual": round(cuota, 2),
            "Pago Interés": round(interes, 2),
            "Pago Capital": round(capital, 2),
            "Saldo Restante": round(saldo, 2),
        })

    return pd.DataFrame(filas)


def calcular_ahorro_compuesto(
    capital_inicial: float,
    aporte_mensual: float,
    tasa_anual: float,
    anios: int,
) -> pd.DataFrame:
    """Proyecta aportes realizados al final de cada mes."""
    tasa_mensual = tasa_anual / 12
    saldo = capital_inicial
    total_aportado = capital_inicial
    filas = []

    for mes in range(1, anios * 12 + 1):
        interes = saldo * tasa_mensual
        saldo += interes + aporte_mensual
        total_aportado += aporte_mensual
        filas.append({
            "Mes": mes,
            "Aporte Mensual": round(aporte_mensual, 2),
            "Interés del Mes": round(interes, 2),
            "Total Aportado": round(total_aportado, 2),
            "Saldo Final": round(saldo, 2),
        })

    return pd.DataFrame(filas)


def crear_reporte_excel(
    datos: pd.DataFrame,
    titulo: str,
    etiquetas: dict[str, str],
    valores: dict[str, object],
    celdas_porcentaje: tuple[str, ...] = (),
) -> bytes:
    """Crea el Excel común de préstamos y ahorros."""
    buffer = io.BytesIO()
    ultima_fila = FILA_ENCABEZADO + len(datos)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        datos.to_excel(writer, sheet_name="Reporte", index=False, startrow=7)
        hoja = writer.sheets["Reporte"]
        fondo_oscuro = PatternFill("solid", fgColor="1A1A1A")
        fondo_resumen = PatternFill("solid", fgColor="252525")
        hoja.sheet_view.showGridLines = False

        for fila in hoja.iter_rows(min_row=1, max_row=7, min_col=1, max_col=5):
            for celda in fila:
                celda.fill = fondo_oscuro

        hoja.merge_cells("A1:E1")
        hoja["A1"] = titulo
        hoja["A1"].font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=16)
        hoja["A1"].alignment = Alignment(horizontal="center", vertical="center")

        hoja.merge_cells("A2:E2")
        hoja["A2"] = "Generado por ZoraEC · Desarrollado por Freddy Beltrán A. (2026)"
        hoja["A2"].font = Font(name="Segoe UI", color="A6A6A6", italic=True, size=10)
        hoja["A2"].alignment = Alignment(horizontal="right", vertical="center")

        for referencia, texto in etiquetas.items():
            celda = hoja[referencia]
            celda.value = texto
            celda.font = Font(name="Segoe UI", color="A6A6A6", size=11)
            celda.alignment = Alignment(horizontal="right", vertical="center")
            celda.fill = fondo_resumen

        for referencia, valor in valores.items():
            celda = hoja[referencia]
            celda.value = valor
            celda.font = Font(name="Segoe UI", color="4DA8DA", bold=True, size=12)
            celda.alignment = Alignment(horizontal="left", vertical="center")
            celda.fill = fondo_resumen
            if referencia not in {"B6"}:
                celda.number_format = (
                    "0.0%" if referencia in celdas_porcentaje else FORMATO_MONEDA
                )

        tabla = Table(displayName="TablaDatos", ref=f"A8:E{ultima_fila}")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleDark1", showRowStripes=True
        )
        hoja.add_table(tabla)

        for fila in hoja.iter_rows(min_row=9, max_row=ultima_fila):
            fila[0].alignment = Alignment(horizontal="center")
            for celda in fila[1:]:
                celda.number_format = FORMATO_MONEDA
                celda.alignment = Alignment(horizontal="right")

        for columna in hoja.columns:
            ancho = max(len(str(celda.value or "")) for celda in columna)
            letra = get_column_letter(columna[0].column)
            hoja.column_dimensions[letra].width = max(ancho + 4, 18)

    return buffer.getvalue()


def generar_excel_prestamo(
    datos: pd.DataFrame, nombre: str, sistema: Sistema
) -> bytes:
    fin = FILA_ENCABEZADO + len(datos)
    return crear_reporte_excel(
        datos,
        f"RESUMEN: {nombre.upper()} ({sistema.upper()})",
        {
            "A4": "Monto del Préstamo:",
            "A5": "Primera Cuota:" if sistema == "Alemán" else "Cuota Fija:",
            "A6": "Plazo:",
            "D4": "Total Intereses:",
            "D5": "Total a Pagar:",
            "D6": "Costo del Crédito:",
        },
        {
            "B4": f"=SUM(D9:D{fin})",
            "B5": "=B9",
            "B6": f'={len(datos)} & " meses"',
            "E4": f"=SUM(C9:C{fin})",
            "E5": f"=SUM(B9:B{fin})",
            "E6": "=(E5/B4)-1",
        },
        ("E6",),
    )


def generar_excel_ahorro(
    datos: pd.DataFrame, capital_inicial: float, aporte_mensual: float
) -> bytes:
    fin = FILA_ENCABEZADO + len(datos)
    return crear_reporte_excel(
        datos,
        "RESUMEN: PLAN DE AHORRO COMPUESTO",
        {
            "A4": "Capital Inicial:",
            "A5": "Aporte Mensual:",
            "A6": "Plazo:",
            "D4": "Total Aportado:",
            "D5": "Total Intereses Ganados:",
            "D6": "Saldo Final:",
        },
        {
            "B4": capital_inicial,
            "B5": aporte_mensual,
            "B6": f'={len(datos)} & " meses"',
            "E4": f"=D{fin}",
            "E5": f"=SUM(C9:C{fin})",
            "E6": f"=E{fin}",
        },
    )


def mostrar_tabla(datos: pd.DataFrame) -> None:
    with st.expander("Ver tabla previa", expanded=False):
        columnas_moneda = [columna for columna in datos.columns if columna != "Mes"]
        formatos = {columna: "${:,.2f}" for columna in columnas_moneda}
        st.dataframe(
            datos.style.format(formatos),
            use_container_width=True,
            hide_index=True,
        )


def aplicar_estilos() -> None:
    """Aplica una identidad visual ligera sin depender de paquetes externos."""
    st.markdown(
        """
        <style>
        .block-container {padding-top: 2rem; padding-bottom: 3rem;}
        [data-testid="stMetric"] {
            background: linear-gradient(145deg, #162033, #111827);
            border: 1px solid #26344d;
            border-radius: 14px;
            padding: 16px;
        }
        [data-testid="stMetricLabel"] {color: #a9b7cc;}
        [data-testid="stMetric"],
        [data-testid="stMetric"] * {box-sizing: border-box; min-width: 0;}
        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] * {
            white-space: normal;
            overflow: visible;
            overflow-wrap: anywhere;
            text-overflow: clip;
        }
        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] * {
            color: #f8fafc;
            font-size: clamp(1.1rem, 2.5vw, 1.65rem);
            line-height: 1.3;
            white-space: normal;
            overflow: visible;
            overflow-wrap: anywhere;
            text-overflow: clip;
        }
        div[data-testid="stForm"] {
            border: 1px solid #26344d;
            border-radius: 16px;
            padding: 1.25rem;
        }
        .decision-box {
            background: linear-gradient(135deg, #132238, #172033);
            border-left: 4px solid #4da8da;
            border-radius: 10px;
            padding: 14px 16px;
            margin: 8px 0 18px 0;
        }
        .decision-box small {color: #a9b7cc;}
        /* El estado nativo de details cambia el signo sin ejecutar Python. */
        [data-testid="stExpander"] details > summary::before {
            content: "+";
            font-size: 1.25rem;
            font-weight: 700;
            flex-shrink: 0;
            width: 1.25rem;
            text-align: center;
        }
        [data-testid="stExpander"] details[open] > summary::before {
            content: "−";
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def tarjeta_contexto(titulo: str, texto: str) -> None:
    st.markdown(
        f'<div class="decision-box"><strong>{titulo}</strong><br>'
        f'<small>{texto}</small></div>',
        unsafe_allow_html=True,
    )


def mostrar_analisis_prestamo(
    datos: pd.DataFrame,
    monto: float,
    ingreso_mensual: float,
    sistema: Sistema,
) -> None:
    primera_cuota = float(datos.iloc[0]["Cuota Mensual"])
    cuota_referencia = primera_cuota if sistema == "Alemán" else float(
        datos["Cuota Mensual"].mean()
    )
    total_intereses = float(datos["Pago Interés"].sum())
    total_pagado = float(datos["Cuota Mensual"].sum())
    costo_porcentual = total_intereses / monto if monto else 0

    st.subheader("Resumen para decidir")
    metrica_1, metrica_2 = st.columns(2)
    metrica_1.metric("Cuota inicial", f"${primera_cuota:,.2f}")
    metrica_2.metric("Intereses totales", f"${total_intereses:,.2f}")
    metrica_3, metrica_4 = st.columns(2)
    metrica_3.metric("Total a pagar", f"${total_pagado:,.2f}")
    metrica_4.metric("Costo financiero", f"{costo_porcentual:.1%}")

    with st.expander("Capacidad de pago", expanded=False):
        if ingreso_mensual > 0:
            carga = cuota_referencia / ingreso_mensual
            st.progress(min(carga, 1.0), text=f"La cuota representa {carga:.1%} del ingreso mensual")
            if carga <= 0.30:
                st.success("La carga estimada está dentro del umbral de referencia del 30 %.")
            elif carga <= 0.40:
                st.warning("La carga supera el 30 %. Conviene revisar gastos y margen de emergencia.")
            else:
                st.error("La carga supera el 40 % del ingreso y podría limitar la liquidez mensual.")
        else:
            st.info("Ingresa tus ingresos mensuales para evaluar la capacidad de pago.")

    with st.expander("Evolución de la deuda", expanded=False):
        grafico_saldo = datos.set_index("Mes")[["Saldo Restante"]]
        st.area_chart(grafico_saldo, color=["#4DA8DA"], use_container_width=True)

    with st.expander("Composición de cada cuota", expanded=False):
        grafico_cuota = datos.set_index("Mes")[["Pago Capital", "Pago Interés"]]
        st.area_chart(
            grafico_cuota,
            color=["#35C48D", "#F4B740"],
            use_container_width=True,
        )


def mostrar_comparacion_sistemas(
    monto: float, tasa: float, plazo: int, sistema_actual: Sistema
) -> None:
    alternativo: Sistema = "Alemán" if sistema_actual == "Francés" else "Francés"
    actual = calcular_amortizacion(monto, tasa, plazo, sistema_actual)
    otro = calcular_amortizacion(monto, tasa, plazo, alternativo)
    comparacion = pd.DataFrame({
        "Sistema": [sistema_actual, alternativo],
        "Primera cuota": [actual.iloc[0]["Cuota Mensual"], otro.iloc[0]["Cuota Mensual"]],
        "Última cuota": [actual.iloc[-1]["Cuota Mensual"], otro.iloc[-1]["Cuota Mensual"]],
        "Intereses totales": [actual["Pago Interés"].sum(), otro["Pago Interés"].sum()],
        "Total a pagar": [actual["Cuota Mensual"].sum(), otro["Cuota Mensual"].sum()],
    })
    ahorro = abs(comparacion.loc[0, "Intereses totales"] - comparacion.loc[1, "Intereses totales"])
    sistema_menor_interes = comparacion.loc[
        comparacion["Intereses totales"].idxmin(), "Sistema"
    ]

    with st.expander("Comparar sistema francés y alemán", expanded=False):
        st.dataframe(
            comparacion.style.format({
                "Primera cuota": "${:,.2f}",
                "Última cuota": "${:,.2f}",
                "Intereses totales": "${:,.2f}",
                "Total a pagar": "${:,.2f}",
            }),
            hide_index=True,
            use_container_width=True,
        )
        st.caption(
            f"El sistema {sistema_menor_interes} genera aproximadamente "
            f"${ahorro:,.2f} menos en intereses para estos parámetros."
        )


def mostrar_analisis_ahorro(
    datos: pd.DataFrame,
    capital_inicial: float,
    meta: float,
) -> None:
    saldo_final = float(datos.iloc[-1]["Saldo Final"])
    total_aportado = float(datos.iloc[-1]["Total Aportado"])
    intereses = saldo_final - total_aportado
    rentabilidad = intereses / total_aportado if total_aportado else 0

    st.subheader("Resumen para decidir")
    metrica_1, metrica_2 = st.columns(2)
    metrica_1.metric("Saldo proyectado", f"${saldo_final:,.2f}")
    metrica_2.metric("Total aportado", f"${total_aportado:,.2f}")
    metrica_3, metrica_4 = st.columns(2)
    metrica_3.metric("Ganancia estimada", f"${intereses:,.2f}")
    metrica_4.metric("Ganancia / aportes", f"{rentabilidad:.1%}")

    with st.expander("Cumplimiento de la meta", expanded=False):
        if meta > 0:
            cumplimiento = saldo_final / meta
            st.progress(min(cumplimiento, 1.0), text=f"Proyección: {cumplimiento:.1%} de la meta")
            diferencia = saldo_final - meta
            if diferencia >= 0:
                st.success(f"La proyección supera la meta por ${diferencia:,.2f}.")
            else:
                st.warning(f"Faltarían ${abs(diferencia):,.2f} para alcanzar la meta.")

        else:
            st.info("Ingresa una meta financiera para evaluar su cumplimiento.")

    with st.expander("Aportes frente al crecimiento acumulado", expanded=False):
        grafico = datos.set_index("Mes")[["Total Aportado", "Saldo Final"]]
        st.line_chart(
            grafico,
            color=["#A9B7CC", "#35C48D"],
            use_container_width=True,
        )

    if capital_inicial == 0 and total_aportado == 0:
        st.info("Agrega capital o aportes mensuales para obtener una proyección útil.")


def mostrar_prestamo(config: ConfiguracionPrestamo) -> None:
    st.title(f"{config.icono} Simulador {config.nombre}")
    st.markdown(config.descripcion)
    tarjeta_contexto(
        config.descripcion,
        "Evalúa la cuota, el costo total y su peso sobre tus ingresos antes de decidir.",
    )

    with st.form(f"formulario_{config.nombre.lower()}"):
        columna_1, columna_2, columna_3 = st.columns(3)
        monto = columna_1.number_input(
            "Monto ($)", min_value=1_000.0, value=config.monto, step=1_000.0
        )
        tasa = columna_2.number_input(
            "Tasa Anual (%)", min_value=0.0, value=config.tasa, step=0.5
        ) / 100
        plazo = columna_3.number_input(
            "Años", min_value=1, value=config.plazo, step=1
        )
        sistema: Sistema = "Francés"
        if config.permite_elegir_sistema:
            sistema = st.radio(
                "Sistema de Amortización:",
                ["Francés", "Alemán"],
                horizontal=True,
            )
        ingreso_mensual = st.number_input(
            "Ingreso mensual disponible ($)",
            min_value=0.0,
            value=2_500.0,
            step=100.0,
            help="Se usa únicamente para estimar qué porcentaje del ingreso representa la cuota.",
        )
        calcular = st.form_submit_button("📊 Calcular y Generar Reporte")

    if not calcular:
        return

    datos = calcular_amortizacion(monto, tasa, plazo, sistema)
    etiqueta = "Cuota Fija Mensual" if sistema == "Francés" else "Primera Cuota"
    st.success(f"**{etiqueta}:** ${datos.iloc[0]['Cuota Mensual']:,.2f}")
    mostrar_analisis_prestamo(datos, monto, ingreso_mensual, sistema)
    if config.permite_elegir_sistema:
        mostrar_comparacion_sistemas(monto, tasa, plazo, sistema)
    mostrar_tabla(datos)
    st.download_button(
        "📥 Descargar Excel",
        data=generar_excel_prestamo(datos, config.nombre, sistema),
        file_name=f"Prestamo_{config.nombre}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def mostrar_ahorro() -> None:
    st.title("💰 Simulador de Ahorro e Inversión")
    st.markdown("Descubre cómo crece tu dinero usando el **interés compuesto**.")
    tarjeta_contexto(
        "Convierte aportes periódicos en una proyección de patrimonio.",
        "Compara lo aportado con la ganancia estimada y verifica si alcanzarías tu meta.",
    )

    with st.form("formulario_ahorro"):
        columna_1, columna_2 = st.columns(2)
        capital = columna_1.number_input(
            "Capital Inicial ($)", min_value=0.0, value=1_000.0, step=500.0
        )
        tasa = columna_1.number_input(
            "Tasa de Rendimiento Anual (%)", min_value=0.0, value=8.0, step=0.5
        ) / 100
        aporte = columna_2.number_input(
            "Aporte Adicional Mensual ($)", min_value=0.0, value=200.0, step=50.0
        )
        plazo = columna_2.number_input(
            "Plazo de Inversión (Años)", min_value=1, value=10, step=1
        )
        meta = st.number_input(
            "Meta financiera ($)",
            min_value=0.0,
            value=40_000.0,
            step=1_000.0,
            help="Se utiliza para medir el cumplimiento de tu objetivo al finalizar el plazo.",
        )
        calcular = st.form_submit_button("📊 Calcular y Generar Reporte")

    if not calcular:
        return

    datos = calcular_ahorro_compuesto(capital, aporte, tasa, plazo)
    columna_1, columna_2 = st.columns(2)
    columna_1.success(
        f"**Saldo Final Proyectado:** ${datos.iloc[-1]['Saldo Final']:,.2f}"
    )
    columna_2.info(
        f"**Total Intereses Ganados:** ${datos['Interés del Mes'].sum():,.2f}"
    )
    mostrar_analisis_ahorro(datos, capital, meta)
    mostrar_tabla(datos)
    st.download_button(
        "📥 Descargar Excel de Ahorro",
        data=generar_excel_ahorro(datos, capital, aporte),
        file_name="Plan_Ahorro_Compuesto.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def main() -> None:
    st.set_page_config(
        page_title="Simulador Financiero 360",
        layout="centered",
        page_icon="🏦",
    )
    aplicar_estilos()
    st.sidebar.title("⚙️ Menú Principal")
    opcion = st.sidebar.radio(
        "Elige el tipo de simulador:", [*PRESTAMOS, "💰 Ahorro / Inversión"]
    )
    st.sidebar.divider()
    st.sidebar.markdown("### 👨‍💻 Creado por ZoraEC")
    st.sidebar.info("**Freddy Beltrán A.**  \nDesarrollador")

    if opcion in PRESTAMOS:
        mostrar_prestamo(PRESTAMOS[opcion])
    else:
        mostrar_ahorro()


if __name__ == "__main__":
    main()
